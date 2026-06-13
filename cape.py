"""
================================================================================
SECTORAL CAPE AND SECTOR RETURNS — FULL ANALYSIS PIPELINE
================================================================================

Author: Fedir Shvets

This single script reproduces every quantitative result reported in Chapters 2
and 3 of the paper, from raw inputs to final tables. It is organised in eight
self-contained stages:

    STAGE 1  Load sectoral earnings (S&P DJI Excel file)
    STAGE 2  Load sector prices (SPDR ETFs) and reduce to quarter-end
    STAGE 3  Load macroeconomic controls (FRED) and build real-rate variables
    STAGE 4  Merge, deflate to real terms, and construct CAPE / rCAPE / returns
    STAGE 5  Estimate the five panel specifications (Table 2.3 / 2.4)
    STAGE 6  Diagnostics: VIF, Hausman, poolability, unit root (Table 2.6)
    STAGE 7  Robustness: raw CAPE, subperiods, sector exclusion, sector-by-sector
    STAGE 8  Backtest (Table 3.1) and out-of-sample forecast (Section 3.5)

--------------------------------------------------------------------------------
METHODOLOGICAL NOTES (read before running)
--------------------------------------------------------------------------------
* CAPE = real price / (40-quarter trailing average of real EPS). The 40-quarter
  (ten-year) window means the first valid CAPE observation is 2017Q4, ten years
  after the earnings series begins in 2008Q1.

* Earnings are S&P index operating EPS; prices are SPDR ETF adjusted-close.
  Because these are denominated on different bases, ABSOLUTE CAPE levels are not
  comparable across sectors. This scaling constant cancels in relative CAPE
  (rCAPE) and is absorbed by sector fixed effects, so all inference uses rCAPE.
  Absolute levels are never interpreted. (See paper, Section 2.2.3.)

* The estimator is the within (fixed-effects) panel estimator. Standard errors
  are two-way clustered by sector and by quarter, robust to heteroscedasticity,
  to the serial correlation induced by overlapping forward returns, and to
  cross-sectional dependence.

* Dependent variable: forward four-quarter (one-year) log total return, in
  percentage points.

--------------------------------------------------------------------------------
REQUIREMENTS
--------------------------------------------------------------------------------
    python >= 3.9
    pandas, numpy, openpyxl, linearmodels, statsmodels, scipy, matplotlib

    pip install pandas numpy openpyxl linearmodels statsmodels scipy matplotlib

--------------------------------------------------------------------------------
INPUT FILES (place in the directory set by DATA_DIR below)
--------------------------------------------------------------------------------
    sp-500-eps-est.xlsx          S&P DJI earnings & estimate report
    XLB_history.csv .. XLY_history.csv   eleven SPDR sector ETF daily histories
    CPILFESL.csv                 FRED core CPI (monthly)
    DFF.csv                      FRED effective federal funds rate (monthly)
    DGS10.csv                    FRED 10-year Treasury yield (monthly)

    ETF tickers used (one CSV each):
    XLB Materials      XLC Comm. Services   XLE Energy        XLF Financials
    XLI Industrials    XLK Info. Technology XLP Cons. Staples XLRE Real Estate
    XLU Utilities      XLV Health Care      XLY Cons. Discr.

--------------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------------
    1. Set DATA_DIR to the folder containing the input files.
    2. Run:  python cape_analysis.py
    3. Output: panel_full.csv.
================================================================================
"""

import os
import re
import warnings

import numpy as np
import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

# ------------------------------------------------------------------ CONFIG ---
DATA_DIR = "."          # <-- set to the folder holding the input files
OUT_DIR = "."           # where CSV outputs and figures are written
MAKE_FIGURES = True     # set False to skip matplotlib figures

# Map S&P sector label (as it appears in the Excel file) -> ETF ticker
SECTOR_MAP = {
    "S&P 500 Consumer Discretionary": "XLY",
    "S&P 500 Consumer Staples":       "XLP",
    "S&P 500 Energy":                 "XLE",
    "S&P 500 Financials ":            "XLF",   # note: trailing space in source
    "S&P 500 Health Care":            "XLV",
    "S&P 500 Industrials":            "XLI",
    "S&P 500 Information Technology":  "XLK",
    "S&P 500 Materials":              "XLB",
    "S&P 500 Communication Services":  "XLC",
    "S&P 500 Utilities":              "XLU",
    "S&P 500 Real Estate (proforma pre-9/19/16)": "XLRE",
}
SECTOR_NAMES = {
    "XLB": "Materials", "XLC": "Comm Services", "XLE": "Energy",
    "XLF": "Financials", "XLI": "Industrials", "XLK": "Info Technology",
    "XLP": "Cons Staples", "XLRE": "Real Estate", "XLU": "Utilities",
    "XLV": "Health Care", "XLY": "Cons Discretionary",
}
WINDOW = 40   # quarters in the CAPE smoothing window (10 years)
HORIZON = 4   # forecast horizon in quarters for the principal dependent variable


# ===========================================================================
# STAGE 1 — SECTORAL EARNINGS
# ===========================================================================
def load_earnings(path):
    """Read quarterly operating EPS per sector from the S&P DJI Excel file.

    The 'SECTOR EPS' sheet has quarter labels ('2008 Q1', ...) in header row 6
    and one row per sector. Returns a long DataFrame: ticker, q (Period), eps.
    """
    wb = load_workbook(path, read_only=True)
    ws = wb["SECTOR EPS"]
    rows = list(ws.iter_rows(min_row=6, max_row=19, values_only=True))
    header = rows[0]
    qcols = [(i, h) for i, h in enumerate(header)
             if isinstance(h, str) and re.match(r"\d{4} Q\d", h)]

    def q_to_period(label):
        y, qq = label.split(" Q")
        return pd.Period(f"{y}Q{qq}", freq="Q")

    records = []
    for r in rows[1:]:
        name = r[0]
        if name in SECTOR_MAP:
            for ci, qlabel in qcols:
                v = r[ci]
                if isinstance(v, (int, float)):
                    records.append({"ticker": SECTOR_MAP[name],
                                    "q": q_to_period(qlabel),
                                    "eps": float(v)})
    eps = pd.DataFrame(records)
    print(f"[STAGE 1] earnings: {len(eps)} rows, "
          f"{eps.ticker.nunique()} sectors, {eps.q.min()}–{eps.q.max()}")
    return eps


# ===========================================================================
# STAGE 2 — SECTOR PRICES (SPDR ETFs, daily -> quarter-end)
# ===========================================================================
def load_prices(data_dir):
    """Read each SPDR ETF daily adjusted-close CSV and take the quarter-end value.

    Quarter-end (last observation of the quarter) is the correct reduction for a
    level/price series; averaging would distort the point-in-time snapshot the
    CAPE numerator requires.
    """
    frames = {}
    for tk in SECTOR_MAP.values():
        df = pd.read_csv(os.path.join(data_dir, f"{tk}_history.csv"),
                         skiprows=3, header=None,
                         names=["Date", "Close", "High", "Low", "Open", "Volume"])
        df["Date"] = pd.to_datetime(df["Date"])
        df = df.dropna(subset=["Close"]).set_index("Date").sort_index()
        frames[tk] = df["Close"].resample("QE").last()   # quarter-end close
    prc = pd.concat(frames, axis=1).stack().rename("price").reset_index()
    prc.columns = ["date", "ticker", "price"]
    prc["q"] = prc["date"].dt.to_period("Q")
    prc = prc[["ticker", "q", "price"]]
    print(f"[STAGE 2] prices: {len(prc)} rows, {prc.q.min()}–{prc.q.max()}")
    return prc


# ===========================================================================
# STAGE 3 — MACROECONOMIC CONTROLS (FRED, monthly -> quarter-end)
# ===========================================================================
def load_macro(data_dir):
    """Build the quarterly macro control set with real interest rates.

        infl     = YoY log change in core CPI (x100)
        infl_10y = trailing 40-quarter mean of infl  (proxy for 10y expectations)
        rffr     = FEDFUNDS - infl                    (real funds rate)
        rdgs10   = DGS10 - infl_10y                   (real 10-year yield)
    """
    def fred(fname, col):
        d = pd.read_csv(os.path.join(data_dir, fname))
        d.columns = ["date", col]
        d["date"] = pd.to_datetime(d["date"])
        d = d.set_index("date").sort_index()
        return d[col].resample("QE").last()

    cpi = fred("CPILFESL.csv", "cpi")
    dff = fred("DFF.csv", "dff")
    dgs = fred("DGS10.csv", "dgs10")
    macro = pd.concat([cpi, dff, dgs], axis=1).reset_index()
    macro["q"] = macro["date"].dt.to_period("Q")
    macro["infl"] = np.log(macro["cpi"] / macro["cpi"].shift(4)) * 100
    macro["infl_10y"] = macro["infl"].rolling(40, min_periods=20).mean()
    macro["rffr"] = macro["dff"] - macro["infl"]
    macro["rdgs10"] = macro["dgs10"] - macro["infl_10y"]
    macro = macro[["q", "cpi", "infl", "rffr", "rdgs10"]]
    print(f"[STAGE 3] macro: {macro.q.min()}–{macro.q.max()}")
    return macro


# ===========================================================================
# STAGE 4 — MERGE, DEFLATE, AND BUILD CAPE / rCAPE / FORWARD RETURNS
# ===========================================================================
def build_panel(eps, prc, macro):
    """Assemble the full panel and construct all model variables.

    Real series use the latest CPI as the base period. CAPE divides the real
    price by the 40-quarter trailing mean of real EPS. rCAPE divides CAPE by the
    sector's own expanding mean. Forward returns are log price ratios h quarters
    ahead. A dummy flags Communication Services from 2018Q4 (the GICS break).
    """
    # IMPORTANT: merge order matters. The 40-quarter earnings average must be
    # computed on the FULL earnings series (which begins in 2008 for every
    # sector, including the proforma Real Estate series), BEFORE restricting to
    # quarters where ETF prices exist. A naive inner merge would truncate the
    # short-history sectors (XLC from 2018, XLRE from 2015) below the 40-quarter
    # window and silently drop them from the panel. We therefore LEFT-merge
    # prices onto the full earnings panel, so earnings rows survive even where
    # prices are absent; CAPE then appears wherever a price exists.
    df = (eps.merge(prc, on=["ticker", "q"], how="left")
              .merge(macro, on="q", how="left")
              .sort_values(["ticker", "q"]).reset_index(drop=True))

    base = macro["cpi"].dropna().iloc[-1]
    df["eps_real"] = df["eps"] * base / df["cpi"]
    df["price_real"] = df["price"] * base / df["cpi"]

    def per_sector(g):
        g = g.sort_values("q").copy()
        # eps_avg40 uses the full earnings history (price not required here)
        g["eps_avg40"] = g["eps_real"].rolling(WINDOW, min_periods=WINDOW).mean()
        # CAPE is defined only where a price exists; NaN price -> NaN CAPE
        g["cape"] = g["price_real"] / g["eps_avg40"]
        g["cape_mean"] = g["cape"].expanding(min_periods=1).mean()
        g["rcape"] = g["cape"] / g["cape_mean"]
        g["ret_fwd4"] = np.log(g["price"].shift(-4) / g["price"])
        g["ret_fwd1"] = np.log(g["price"].shift(-1) / g["price"])
        g["ret_fwd12"] = np.log(g["price"].shift(-12) / g["price"])
        g["ret_next"] = g["price"].shift(-1) / g["price"] - 1   # for backtest
        return g

    df = pd.concat([per_sector(g) for _, g in df.groupby("ticker")],
                   ignore_index=True)
    df["dummy_comm"] = ((df["ticker"] == "XLC") &
                        (df["q"] >= pd.Period("2018Q4", freq="Q"))).astype(int)

    panel = df.dropna(subset=["cape"]).copy()
    print(f"[STAGE 4] CAPE defined {panel.q.min()}–{panel.q.max()}, "
          f"{len(panel)} obs, {panel.ticker.nunique()} sectors")
    df.to_csv(os.path.join(OUT_DIR, "panel_full.csv"), index=False)
    return df


# ===========================================================================
# STAGE 5 — PANEL ESTIMATION (Tables 2.3 and 2.4)
# ===========================================================================
def estimate(df):
    from linearmodels.panel import PanelOLS, PooledOLS

    d = df.dropna(subset=["ret_fwd4", "rcape", "rffr", "rdgs10", "infl"]).copy()
    d["ret_fwd4"] = d["ret_fwd4"] * 100      # percentage points
    d["qt"] = d["q"].dt.to_timestamp()
    dd = d.set_index(["ticker", "qt"]).sort_index()

    clu = dict(cov_type="clustered", cluster_entity=True, cluster_time=True)

    print("\n" + "=" * 78)
    print("TABLE 2.3 — FIVE PANEL SPECIFICATIONS  (dep = forward 1yr return, %)")
    print("=" * 78)
    print(f"{'Specification':<30}{'beta':>9}{'SE':>9}{'t':>7}{'p':>8}{'R2w':>8}{'N':>6}")
    print("-" * 78)

    out = {}
    # (1) Pooled OLS
    m = PooledOLS(dd["ret_fwd4"],
                  dd[["rcape", "rffr", "rdgs10", "infl"]].assign(const=1)).fit(**clu)
    out["pooled"] = m
    _row("(1) Pooled OLS", m, "rcape", m.rsquared)
    # (2) Sector FE only
    m = PanelOLS(dd["ret_fwd4"], dd[["rcape"]], entity_effects=True).fit(**clu)
    out["fe"] = m
    _row("(2) Sector FE", m, "rcape", m.rsquared_within)
    # (3) Sector FE + controls  [PRINCIPAL]
    m = PanelOLS(dd["ret_fwd4"], dd[["rcape", "rffr", "rdgs10", "infl"]],
                 entity_effects=True).fit(**clu)
    out["principal"] = m
    _row("(3) Sector FE + controls", m, "rcape", m.rsquared_within)
    # (4) Sector FE + controls, h = 1 quarter
    d1 = df.dropna(subset=["ret_fwd1", "rcape", "rffr", "rdgs10", "infl"]).copy()
    d1["ret_fwd1"] = d1["ret_fwd1"] * 100
    d1["qt"] = d1["q"].dt.to_timestamp()
    ix1 = d1.set_index(["ticker", "qt"]).sort_index()
    m = PanelOLS(ix1["ret_fwd1"], ix1[["rcape", "rffr", "rdgs10", "infl"]],
                 entity_effects=True).fit(**clu)
    out["h1"] = m
    _row("(4) Sector FE + controls h=1", m, "rcape", m.rsquared_within)
    # (5) Sector + Time FE (controls drop out: collinear with time effects)
    m = PanelOLS(dd["ret_fwd4"], dd[["rcape"]],
                 entity_effects=True, time_effects=True).fit(**clu)
    out["twoway"] = m
    _row("(5) Sector + Time FE", m, "rcape", m.rsquared_within)
    print("=" * 78)

    print("\nTABLE 2.4 — PRINCIPAL SPECIFICATION (3), FULL COEFFICIENTS")
    print("-" * 60)
    p = out["principal"]
    for v in ["rcape", "rffr", "rdgs10", "infl"]:
        print(f"  {v:<10} coef={p.params[v]:8.3f}  SE={p.std_errors[v]:6.3f}"
              f"  t={p.tstats[v]:6.2f}  p={p.pvalues[v]:.3f}")
    print(f"  N={int(p.nobs)}, within-R2={p.rsquared_within:.3f}, "
          f"F p-value={p.f_statistic.pval:.4g}")

    return out, dd


def _row(name, res, key, r2):
    b = res.params[key]; se = res.std_errors[key]
    t = res.tstats[key]; pv = res.pvalues[key]
    star = "***" if pv < .01 else "**" if pv < .05 else "*" if pv < .1 else ""
    print(f"{name:<30}{b:>9.2f}{se:>9.2f}{t:>7.2f}{pv:>8.3f}{r2:>8.3f}"
          f"{int(res.nobs):>6}  {star}")


# ===========================================================================
# STAGE 6 — DIAGNOSTICS (VIF, Hausman, poolability, unit root)  Table 2.6
# ===========================================================================
def diagnostics(df, models, dd):
    from linearmodels.panel import RandomEffects
    from statsmodels.stats.outliers_influence import variance_inflation_factor
    from statsmodels.tsa.stattools import adfuller
    from scipy.stats import chi2

    print("\n" + "=" * 60)
    print("TABLE 2.6 — VARIANCE INFLATION FACTORS")
    print("=" * 60)
    d = df.dropna(subset=["ret_fwd4", "rcape", "rffr", "rdgs10", "infl"])
    X = d[["rcape", "rffr", "rdgs10", "infl"]].assign(const=1.0)
    for i, c in enumerate(["rcape", "rffr", "rdgs10", "infl"]):
        print(f"  {c:<8} VIF = {variance_inflation_factor(X.values, i):5.2f}")

    # Hausman (FE vs RE)
    fe = models["principal"]
    re = RandomEffects(dd["ret_fwd4"],
                       dd[["rcape", "rffr", "rdgs10", "infl"]].assign(const=1)).fit()
    cols = ["rcape", "rffr", "rdgs10", "infl"]
    diff = fe.params[cols].values - re.params[cols].values
    vdiff = fe.cov.loc[cols, cols].values - re.cov.loc[cols, cols].values
    stat = float(diff @ np.linalg.pinv(vdiff) @ diff)
    print(f"\nHausman (FE vs RE): chi2={stat:.2f}, "
          f"p={1 - chi2.cdf(stat, len(cols)):.3f}")
    print(f"Poolability F-test: F={fe.f_pooled.stat:.2f}, p={fe.f_pooled.pval:.4f}")

    # Panel unit root (per-sector ADF on CAPE)
    rej = tot = 0
    for tk, g in df.dropna(subset=["cape"]).groupby("ticker"):
        s = g.sort_values("q")["cape"].values
        if len(s) > 12:
            tot += 1
            try:
                if adfuller(s, maxlag=2, autolag=None)[1] < 0.10:
                    rej += 1
            except Exception:
                pass
    print(f"Unit-root (ADF) on CAPE: {rej}/{tot} sectors reject at 10%")


# ===========================================================================
# STAGE 7 — ROBUSTNESS (Tables 2.5, 2.7, 2.8)
# ===========================================================================
def robustness(df):
    from linearmodels.panel import PanelOLS
    import statsmodels.api as sm

    d = df.copy()
    d["ret_fwd4"] = d["ret_fwd4"] * 100
    d["qt"] = d["q"].dt.to_timestamp()
    clu = dict(cov_type="clustered", cluster_entity=True, cluster_time=True)

    def fit(data, reg):
        sub = data.dropna(subset=["ret_fwd4", reg, "rffr", "rdgs10", "infl"])
        ix = sub.set_index(["ticker", "qt"]).sort_index()
        return PanelOLS(ix["ret_fwd4"], ix[[reg, "rffr", "rdgs10", "infl"]],
                        entity_effects=True).fit(**clu)

    print("\n" + "=" * 60)
    print("TABLE 2.8 — RAW vs RELATIVE CAPE, AND SECTOR EXCLUSION")
    print("=" * 60)
    for lbl, reg, data in [("Relative CAPE", "rcape", d),
                           ("Raw CAPE", "cape", d),
                           ("Excl. XLC (rcape)", "rcape", d[d.ticker != "XLC"])]:
        m = fit(data, reg)
        print(f"  {lbl:<20} beta={m.params[reg]:8.2f}  "
              f"p={m.pvalues[reg]:.3f}  N={int(m.nobs)}")

    print("\nTABLE 2.7 — SUBPERIODS (rcape, sector FE + controls)")
    subs = [("Pre-COVID 2017Q4-2019Q4", "2017-10-01", "2019-12-31"),
            ("COVID 2020Q1-2021Q4",     "2020-01-01", "2021-12-31"),
            ("Tightening 2022Q1-2024Q2","2022-01-01", "2024-06-30")]
    for lbl, a, b in subs:
        sub = d[(d.qt >= a) & (d.qt <= b)]
        try:
            m = fit(sub, "rcape")
            print(f"  {lbl:<26} beta={m.params['rcape']:8.2f}  "
                  f"p={m.pvalues['rcape']:.3f}  N={int(m.nobs)}")
        except Exception:
            print(f"  {lbl:<26} (insufficient observations)")

    print("\nTABLE 2.5 — SECTOR-BY-SECTOR (univariate OLS)")
    rows = []
    for tk, g in d.groupby("ticker"):
        g = g.dropna(subset=["ret_fwd4", "rcape"])
        if len(g) > 10:
            X = sm.add_constant(g["rcape"])
            m = sm.OLS(g["ret_fwd4"], X).fit()
            rows.append((tk, m.params["rcape"], m.tvalues["rcape"],
                         m.pvalues["rcape"]))
    for tk, b, t, p in sorted(rows, key=lambda r: r[1]):
        star = "***" if p < .01 else "**" if p < .05 else "*" if p < .1 else ""
        print(f"  {tk:<5} {SECTOR_NAMES[tk]:<18} beta={b:8.1f}  "
              f"t={t:6.2f}  p={p:.3f}  {star}")


# ===========================================================================
# STAGE 8 — BACKTEST (Table 3.1) AND OUT-OF-SAMPLE FORECAST (Section 3.5)
# ===========================================================================
def backtest(df):
    from scipy import stats as st

    d = df.dropna(subset=["rcape", "ret_next"]).copy()
    quarters = sorted(d["q"].unique())
    strat, bench = [], []
    for q in quarters:
        g = d[d["q"] == q]
        if g["ticker"].nunique() < 8:
            continue
        g = g.sort_values("rcape")
        k = max(3, len(g) // 3)               # cheapest tercile
        strat.append(g.head(k)["ret_next"].mean())
        bench.append(g["ret_next"].mean())     # equal-weight benchmark
    strat, bench = np.array(strat), np.array(bench)

    def stats(r):
        ar = (1 + r.mean()) ** 4 - 1
        av = r.std(ddof=1) * np.sqrt(4)
        eq = np.cumprod(1 + r); dd = (eq - np.maximum.accumulate(eq)) / np.maximum.accumulate(eq)
        return ar, av, (ar / av if av else np.nan), np.prod(1 + r) - 1, dd.min()

    print("\n" + "=" * 60)
    print("TABLE 3.1 — BACKTEST: CHEAP-TERCILE STRATEGY vs EQUAL-WEIGHT")
    print("=" * 60)
    for name, r in [("Strategy", strat), ("Benchmark", bench)]:
        ar, av, sh, cu, md = stats(r)
        print(f"  {name:<10} annRet={ar*100:5.1f}%  annVol={av*100:5.1f}%  "
              f"Sharpe={sh:.2f}  cum={cu*100:5.1f}%  maxDD={md*100:5.1f}%")
    diff = strat - bench
    t, p = st.ttest_1samp(diff, 0)
    print(f"  Outperformance: {diff.mean()*100:.2f}%/q  t={t:.2f}  p={p:.3f}  "
          f"win-rate={(diff > 0).mean()*100:.0f}%")


def out_of_sample(df):
    from linearmodels.panel import PanelOLS
    from scipy import stats as st

    d = df.dropna(subset=["ret_fwd4", "rcape", "rffr", "rdgs10", "infl"]).copy()
    d["ret_fwd4"] = d["ret_fwd4"] * 100
    d["qt"] = d["q"].dt.to_timestamp()
    quarters = sorted(d["q"].unique())
    split = quarters[len(quarters) // 2]      # expanding window starts midway

    preds, bench, actual = [], [], []
    for q in [x for x in quarters if x >= split]:
        train = d[d["q"] < q]; test = d[d["q"] == q]
        if train["ticker"].nunique() < 8 or len(test) == 0:
            continue
        ix = train.set_index(["ticker", "qt"]).sort_index()
        m = PanelOLS(ix["ret_fwd4"], ix[["rcape", "rffr", "rdgs10", "infl"]],
                     entity_effects=True).fit()
        b = m.params
        intercept = (train["ret_fwd4"].mean()
                     - sum(b[v] * train[v].mean()
                           for v in ["rcape", "rffr", "rdgs10", "infl"]))
        for _, row in test.iterrows():
            yhat = intercept + sum(b[v] * row[v]
                                   for v in ["rcape", "rffr", "rdgs10", "infl"])
            preds.append(yhat); bench.append(train["ret_fwd4"].mean())
            actual.append(row["ret_fwd4"])

    preds, bench, actual = map(np.array, (preds, bench, actual))
    mse_m = np.mean((actual - preds) ** 2)
    mse_b = np.mean((actual - bench) ** 2)
    t, p = st.ttest_1samp((actual - bench) ** 2 - (actual - preds) ** 2, 0)
    print("\n" + "=" * 60)
    print("SECTION 3.5 — OUT-OF-SAMPLE FORECAST (expanding window)")
    print("=" * 60)
    print(f"  OOS obs={len(actual)}  model RMSE={np.sqrt(mse_m):.2f}  "
          f"benchmark RMSE={np.sqrt(mse_b):.2f}")
    print(f"  OOS R-squared = {1 - mse_m/mse_b:.4f}  "
          f"(negative => worse than historical mean)")
    print(f"  Diff-in-sq-errors t={t:.2f}, p={p:.3f}")


# ===========================================================================
# MAIN
# ===========================================================================
def main():
    print("=" * 78)
    print("SECTORAL CAPE — FULL ANALYSIS PIPELINE")
    print("=" * 78)
    eps = load_earnings(os.path.join(DATA_DIR, "sp-500-eps-est.xlsx"))
    prc = load_prices(DATA_DIR)
    macro = load_macro(DATA_DIR)
    df = build_panel(eps, prc, macro)

    models, dd = estimate(df)
    diagnostics(df, models, dd)
    robustness(df)
    backtest(df)
    out_of_sample(df)

    print("\n[done] panel_full.csv written to", os.path.abspath(OUT_DIR))


if __name__ == "__main__":
    main()
